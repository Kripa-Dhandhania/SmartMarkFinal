"""
Unified Smart Attendance System
Integrates Module 3 (Analytics Dashboard) + SmartMark (Face Recognition + OTP)
"""

from flask import Flask, render_template, redirect, url_for, request, session, flash, Response, jsonify
import os
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, datetime, timedelta
import math
import hashlib
import base64
import time
from PIL import Image
import numpy as np
import cv2
from werkzeug.security import generate_password_hash, check_password_hash

# ── SmartMark imports ──────────────────────────────────────────────────────────
from face_manager import capture_face, save_face, train_recognizer, recognize_face, check_face_enrolled
from otp_manager import generate_otp, send_otp_email, send_otp_email_async, validate_otp
from notification_manager import send_low_attendance_alert, send_weekly_digest
from database import (
    init_db,
    get_or_create_student, get_student, get_student_email, update_student_email, add_student,
    check_face_exists_in_db, get_connection,
    create_session, close_session, get_active_sessions, get_session,
    check_attendance_for_session, mark_attendance,
    get_attendance_records,
    get_analytics_stats, get_active_session_stats, get_daily_trend, get_defaulters,
    get_all_students, get_student_attendance_summary,
    get_verification_method_stats,
    get_student_subject_summary,
    get_or_create_teacher, get_teacher, get_teacher_subjects, add_teacher, add_teacher_subject,
    create_leave_request, get_leave_requests, update_leave_status
)

# ── App setup ──────────────────────────────────────────────────────────────────
app = Flask(__name__,
    static_folder=os.path.join(os.path.dirname(__file__), "static"),
    static_url_path="/static"
)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "smartmark_unified_secret_key_2024")

# Global recognizer state
recognizer = None
label_map = None

# Initialise DB on startup
init_db()

# ── Default Profile ──────────────────────────────────────────────────────────
DEFAULT_TEACHER = {
    "name": "Dr. R. Sridevi",
    "department": "Department of Computer Science",
    "email": "sridevir@christuniversity.in"
}

CLASSROOM_LAT = float(os.environ.get("CLASSROOM_LAT", 12.933416887384725))
CLASSROOM_LON = float(os.environ.get("CLASSROOM_LON", 77.60607102136017))
CLASSROOM_RADIUS = float(os.environ.get("CLASSROOM_RADIUS", 20.0))


SUBJECTS = [
    {"name": "Software Development Project",     "class_info": "MCA - Sem 3"},
    {"name": "Mobile Application Development",    "class_info": "MCA - Sem 3"},
    {"name": "Cognitive Psychology",             "class_info": "MCA - Sem 3"},
    {"name": "Go Programming",                  "class_info": "MCA - Sem 3"},
    {"name": "Data Communications and Networks", "class_info": "MCA - Sem 3"},
]

# ── Helper Functions ──────────────────────────────────────────────────────────
def haversine_distance(lat1, lon1, lat2, lon2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees) in meters.
    """
    R = 6371000  # Radius of earth in meters
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)

    a = math.sin(dphi / 2)**2 + \
        math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    return R * c

def get_dynamic_otp(session_id):
    """Generate a dynamic 6-digit numeric OTP that changes every 10 seconds."""
    secret = app.secret_key
    # 10-second slots
    time_slot = int(time.time() // 10)
    token_input = f"{session_id}{secret}{time_slot}"
    hash_val = int(hashlib.sha256(token_input.encode()).hexdigest(), 16)
    return str(hash_val % 1000000).zfill(6)

def verify_otp(session_id, token):
    """Verify if the provided 6-digit OTP is valid for the given session_id (allows +/- 1 slot)."""
    if not token:
        return False
    current_slot = int(time.time() // 10)
    # Check current and previous 10s slot for network lag buffer
    for slot in [current_slot, current_slot - 1]:
        token_input = f"{session_id}{app.secret_key}{slot}"
        hash_val = int(hashlib.sha256(token_input.encode()).hexdigest(), 16)
        expected = str(hash_val % 1000000).zfill(6)
        if str(token) == expected:
            return True
    return False

def decode_base64_image(data_url):
    """Convert a browser-side data URL (Base64) to an OpenCV grayscale image."""
    try:
        if ',' in data_url:
            data_url = data_url.split(',')[1]
        
        image_data = base64.b64decode(data_url)
        image = Image.open(io.BytesIO(image_data))
        
        # Convert to OpenCV format (Grayscale)
        frame = np.array(image.convert('RGB'))
        gray = cv2.cvtColor(frame, cv2.COLOR_RGB2GRAY)
        
        # Detect face in the captured frame
        from face_manager import get_face_cascade
        face_cascade = get_face_cascade()
        faces = face_cascade.detectMultiScale(gray, 1.1, 5, minSize=(100, 100))
        
        if len(faces) > 0:
            (x, y, w, h) = faces[0]
            face_img = gray[y:y+h, x:x+w]
            face_img = cv2.resize(face_img, (200, 200))
            face_img = cv2.equalizeHist(face_img)
            return face_img
        
        # Fallback: if no face detected in specialized crop, return centered crop or original
        return cv2.resize(gray, (200, 200))
    except Exception as e:
        print(f"[ERROR] decode_base64_image: {e}")
        return None

# ══════════════════════════════════════════════════════════════════════════════
#  TEACHER ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def home():
    if "teacher_id" in session:
        return redirect(url_for("teacher_dashboard"))
    return redirect(url_for("teacher_login"))


@app.route("/teacher/login", methods=["GET", "POST"])
def teacher_login():
    if request.method == "POST":
        password   = request.form.get("password", "").strip()

        teacher_id = request.form.get("teacher_id", "").strip()
        if not teacher_id or not password:
            flash("Teacher ID and Password are required.", "error")
            return redirect(url_for("teacher_login"))

        teacher = get_teacher(teacher_id)
        if teacher and teacher.get('password_hash'):
            if check_password_hash(teacher['password_hash'], password):
                session["teacher_id"] = teacher_id
                flash(f"Welcome, {teacher['name']}!", "success")
                return redirect(url_for("teacher_dashboard"))
            else:
                flash("Invalid password.", "error")
        elif teacher:
            # Legacy account without password
            flash("This account requires a password update. Please register again with a password.", "warning")
        else:
            flash("Teacher ID not found. Please register.", "error")

        return redirect(url_for("teacher_login"))

    return render_template("teacher_login.html")


@app.route("/teacher/register", methods=["GET", "POST"])
def teacher_register():
    if request.method == "POST":
        teacher_id = request.form.get("teacher_id", "").strip()
        name       = request.form.get("name", "").strip()
        email      = request.form.get("email", "").strip()
        password     = request.form.get("password", "").strip()
        subjects_str = request.form.get("subjects", "").strip()

        if not teacher_id or not name or not email or not password:
            flash("Teacher ID, Name, Email, and Password are required.", "error")
            return redirect(url_for("teacher_register"))

        teacher = get_teacher(teacher_id)
        if teacher and teacher.get('password_hash'):
            flash("Teacher ID already registered. Please login.", "warning")
            return redirect(url_for("teacher_login"))

        dept         = request.form.get("department", "").strip()
        password_hash = generate_password_hash(password)
        if add_teacher(teacher_id, name, email, dept):
            from database import update_teacher_password
            update_teacher_password(teacher_id, password_hash)
            
            # Process subjects
            if subjects_str:
                subject_list = [s.strip() for s in subjects_str.split(",") if s.strip()]
                for sname in subject_list:
                    add_teacher_subject(teacher_id, sname)
            
            session["teacher_id"] = teacher_id
            flash(f"Welcome, {name}! Your registration was successful.", "success")
            return redirect(url_for("teacher_dashboard"))
        else:
            flash("Registration failed. Please try again.", "error")
            return redirect(url_for("teacher_register"))

    return render_template("teacher_login.html", register=True)


@app.route("/teacher/logout")
def teacher_logout():
    session.pop("teacher_id", None)
    flash("Logged out from teacher portal.", "info")
    return redirect(url_for("teacher_login"))


@app.route("/teacher")
def teacher_dashboard():
    if "teacher_id" not in session:
        return redirect(url_for("teacher_login"))
        
    teacher_id = session["teacher_id"]
    teacher = get_teacher(teacher_id) or DEFAULT_TEACHER
    
    # Get subjects assigned to this teacher
    teacher_subjects = get_teacher_subjects(teacher_id)
    subject_names = [s['name'] for s in teacher_subjects]
    
    stats = get_analytics_stats(subjects=subject_names)
    active_sessions = get_active_sessions(subjects=subject_names)
    # Determine current hour based on time
    now_hour = datetime.now().hour
    current_hour = 1
    if now_hour >= 9: current_hour = 2
    if now_hour >= 11: current_hour = 3
    if now_hour >= 12: current_hour = 4
    if now_hour >= 13: current_hour = 5
    
    from database import get_pending_verifications
    pending_verifications = get_pending_verifications(teacher_id)

    return render_template(
        "teacher_dashboard.html",
        teacher=teacher,
        subjects=teacher_subjects,
        active_sessions=active_sessions,
        stats=stats,
        current_hour=current_hour,
        hour_labels=HOUR_LABELS,
        pending_verifications=pending_verifications
    )



@app.route("/teacher/create-session", methods=["POST"])
def teacher_create_session():
    if "teacher_id" not in session:
        return redirect(url_for("teacher_login"))
        
    subject = request.form.get("subject", "").strip()
    hour    = request.form.get("hour", "").strip()

    if not subject or not hour:
        flash("Subject and Hour are required.", "error")
        return redirect(url_for("teacher_dashboard"))

    try:
        hour_int = int(hour)
    except ValueError:
        flash("Hour must be a number.", "error")
        return redirect(url_for("teacher_dashboard"))

    teacher_id = session.get("teacher_id")
    duration = request.form.get("duration", "5")
    lat = request.form.get("latitude")
    lon = request.form.get("longitude")
    
    try:
        duration_int = int(duration)
    except ValueError:
        duration_int = 5
        
    create_session(subject, hour_int, teacher_id, ttl_minutes=duration_int, lat=lat, lon=lon)
    flash(f"Attendance enabled: {subject} – Hour {hour_int} for {duration_int} mins", "success")
    return redirect(url_for("teacher_dashboard"))



@app.route("/teacher/close-session/<int:session_id>", methods=["POST"])
def teacher_close_session(session_id):
    if "teacher_id" not in session:
        return redirect(url_for("teacher_login"))
    close_session(session_id)
    flash("Session closed.", "success")
    return redirect(url_for("teacher_dashboard"))


@app.route("/teacher/api/close-session/<int:session_id>", methods=["POST"])
def teacher_api_close_session(session_id):
    """AJAX endpoint – used by the JS countdown timer to auto-close a session."""
    from flask import jsonify
    close_session(session_id)
    return jsonify({"status": "closed", "session_id": session_id})


@app.route("/teacher/api/extend-session/<int:session_id>", methods=["POST"])
def teacher_api_extend_session(session_id):
    """AJAX endpoint to increase the session timer."""
    from flask import jsonify
    from database import extend_session
    minutes = request.json.get("minutes", 2)
    success = extend_session(session_id, minutes)
    return jsonify({"success": success, "session_id": session_id, "extended_by": minutes})


@app.route("/teacher/api/manual-mark-attendance", methods=["POST"])
def teacher_api_manual_mark_attendance():
    """AJAX endpoint for teachers to manually mark a student present."""
    if "teacher_id" not in session:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
        
    data = request.json
    student_id = data.get("student_id")
    session_id = data.get("session_id")
    
    if not student_id or not session_id:
        return jsonify({"success": False, "error": "Missing student_id or session_id"}), 400
        
    from database import mark_attendance, get_session
    
    # Check if student exists
    student = get_student(student_id)
    if not student:
        return jsonify({"success": False, "error": "Student not found"}), 404
        
    # Mark attendance
    success = mark_attendance(student_id, session_id, auth_method="Manual")
    
    if success:
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "error": "Attendance already marked or failed"}), 400


@app.route("/api/students/search")
def api_students_search():
    """Search for students by name or ID (partial match)."""
    if "teacher_id" not in session:
        return jsonify({"success": False, "error": "Unauthorized"}), 403
        
    query = request.args.get("q", "").lower().strip()
    if not query:
        return jsonify([])
        
    all_students = get_all_students()
    results = [
        {"id": s["id"], "name": s["name"]}
        for s in all_students
        if query in s["id"].lower() or query in s["name"].lower()
    ]
    return jsonify(results[:10]) # Limit to 10 results


@app.route("/teacher/approve-verification/<int:attendance_id>", methods=["POST"])
def teacher_approve_verification(attendance_id):
    if "teacher_id" not in session:
        return redirect(url_for("teacher_login"))
    
    from database import approve_attendance, get_session, get_connection
    
    # Security: Verify this attendance record belongs to a session of the logged-in teacher
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT session_id FROM attendance WHERE id = ?", (attendance_id,))
        att = cursor.fetchone()
        if not att:
            flash("Attendance record not found.", "error")
            return redirect(url_for("teacher_dashboard"))
        
        sess = get_session(att['session_id'])
        if not sess or sess['teacher_id'] != session['teacher_id']:
            flash("Unauthorized: You can only verify students for your own sessions.", "error")
            return redirect(url_for("teacher_dashboard"))

    if approve_attendance(attendance_id):
        flash("Student verified successfully!", "success")
    else:
        flash("Failed to verify student.", "error")
    return redirect(url_for("teacher_dashboard"))


@app.route("/teacher/reject-verification/<int:attendance_id>", methods=["POST"])
def teacher_reject_verification(attendance_id):
    if "teacher_id" not in session:
        return redirect(url_for("teacher_login"))
    
    from database import reject_attendance, get_session, get_connection
    
    # Security: Verify this attendance record belongs to a session of the logged-in teacher
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT session_id FROM attendance WHERE id = ?", (attendance_id,))
        att = cursor.fetchone()
        if not att:
            flash("Attendance record not found.", "error")
            return redirect(url_for("teacher_dashboard"))
        
        sess = get_session(att['session_id'])
        if not sess or sess['teacher_id'] != session['teacher_id']:
            flash("Unauthorized: You can only verify students for your own sessions.", "error")
            return redirect(url_for("teacher_dashboard"))

    if reject_attendance(attendance_id):
        flash("Verification request rejected.", "warning")
    else:
        flash("Failed to reject request.", "error")
    return redirect(url_for("teacher_dashboard"))




# ══════════════════════════════════════════════════════════════════════════════
#  ANALYTICS ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/analytics")
def analytics():
    if "teacher_id" not in session:
        return redirect(url_for("teacher_login"))
        
    teacher_id = session.get("teacher_id")
    teacher = get_teacher(teacher_id) or DEFAULT_TEACHER
    
    # Get subjects assigned to this teacher
    teacher_subjects = get_teacher_subjects(teacher_id) if teacher_id else []
    subject_names = [s['name'] for s in teacher_subjects]
    
    stats        = get_analytics_stats(subjects=subject_names)
    daily_trend  = get_daily_trend(days=7, subjects=subject_names)
    defaulters   = get_defaulters(threshold=75, subjects=subject_names)
    method_stats = get_verification_method_stats(subjects=subject_names)
    students     = get_student_attendance_summary(subjects=subject_names)
    return render_template(
        "analytics.html",
        teacher=teacher,
        stats=stats,
        daily_trend=daily_trend,
        defaulters=defaulters,
        method_stats=method_stats,
        students=students
    )


@app.route("/api/analytics/live")
def api_analytics_live():
    if "teacher_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    teacher_id = session.get("teacher_id")
    teacher_subjects = get_teacher_subjects(teacher_id) if teacher_id else []
    subject_names = [s['name'] for s in teacher_subjects]
    
    # Check for active session
    active_session_stats = get_active_session_stats(teacher_id)
    
    stats = get_analytics_stats(subjects=subject_names)
    daily_trend = get_daily_trend(days=7, subjects=subject_names)
    method_stats = get_verification_method_stats(subjects=subject_names)
    defaulters = get_defaulters(threshold=75, subjects=subject_names)
    students = get_student_attendance_summary(subjects=subject_names)
    
    return jsonify({
        "stats": stats,
        "active_session": active_session_stats,
        "daily_trend": daily_trend,
        "method_stats": method_stats,
        "defaulters": defaulters,
        "students": students
    })


@app.route("/teacher/send-alerts", methods=["POST"])
def teacher_send_alerts():
    if "teacher_id" not in session:
        return redirect(url_for("teacher_login"))
        
    teacher_id = session.get("teacher_id")
    teacher_subjects = get_teacher_subjects(teacher_id) if teacher_id else []
    subject_names = [s['name'] for s in teacher_subjects] if teacher_subjects else None

    defaulters = get_defaulters(threshold=75, subjects=subject_names)
    sent_count = 0
    
    for d in defaulters:
        if d.get("email"):
            send_low_attendance_alert(d["name"], d["email"], d["percentage"])
            sent_count += 1
            
    flash(f"Low-attendance alerts sent to {sent_count} students.", "success")
    return redirect(url_for("analytics"))


@app.route("/teacher/send-weekly-digest", methods=["POST"])
def teacher_send_weekly_digest():
    if "teacher_id" not in session:
        return redirect(url_for("teacher_login"))
        
    teacher_id = session["teacher_id"]
    teacher = get_teacher(teacher_id) or DEFAULT_TEACHER
    
    teacher_subjects = get_teacher_subjects(teacher_id)
    subject_names = [s['name'] for s in teacher_subjects]

    stats = get_analytics_stats(subjects=subject_names)
    # Add a date range for the digest
    stats["date_range"] = f"{(date.today() - timedelta(days=7)).isoformat()} to {date.today().isoformat()}"
    stats["avg_attendance"] = stats["attendance_pct"] # Simplified for now
    stats["sessions_count"] = stats["sessions_today"] * 5 # Representative estimate
    
    defaulters = get_defaulters(threshold=75, subjects=subject_names)
    
    send_weekly_digest(teacher["name"], teacher["email"], stats, defaulters)
    
    flash("Weekly digest summary sent to your email.", "success")
    return redirect(url_for("analytics"))


# ══════════════════════════════════════════════════════════════════════════════
#  ATTENDANCE REPORT ROUTES  (real DB data)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/attendance-report")
def attendance_report():
    selected_date = request.args.get("date", date.today().isoformat())
    student_filter = request.args.get("student_id", "")
    hour_filter = request.args.get("hour", "all")
    mismatch_filter = request.args.get("status", "")  # Use "status" for mismatch filters

    teacher_id = session.get("teacher_id")
    teacher_subjects = get_teacher_subjects(teacher_id) if teacher_id else []
    subject_names = [s['name'] for s in teacher_subjects]

    all_students = get_all_students()
    
    # Base records for the day and teacher's subjects
    records = get_attendance_records_for_report(selected_date, student_filter, subjects=subject_names)

    # Filter by hour slot
    if hour_filter != "all":
        try:
            h = int(hour_filter)
            records = [r for r in records if str(r.get("hour", "")) == str(h)]
        except ValueError:
            pass

    # Apply special mismatch filters
    if mismatch_filter:
        # Get all records for this date to perform cross-hour analysis
        day_records = get_attendance_records_for_report(selected_date, "", subjects=subject_names)
        
        student_hours = {}
        for r in day_records:
            sid = r.get("student_id")
            h = str(r.get("hour", ""))
            if sid not in student_hours:
                student_hours[sid] = set()
            if h:
                student_hours[sid].add(h)
        
        target_student_ids = []
        for s in all_students:
            sid = s["id"]
            hours = student_hours.get(sid, set())
            has_h1 = "1" in hours
            has_h2 = "2" in hours
            
            if mismatch_filter == "missed1_att2":
                if not has_h1 and has_h2:
                    target_student_ids.append(sid)
            elif mismatch_filter == "att1_missed2":
                if has_h1 and not has_h2:
                    target_student_ids.append(sid)
                    
        records = [r for r in records if r.get("student_id") in target_student_ids]


    teacher_id = session.get("teacher_id")
    teacher = get_teacher(teacher_id) if teacher_id else DEFAULT_TEACHER

    return render_template(
        "attendance_report.html",
        teacher=teacher,
        records=records,
        selected_date=selected_date,
        student_filter=student_filter,
        hour_filter=hour_filter,
        mismatch_filter=mismatch_filter,
        all_students=all_students
    )


@app.route("/download-report")
def download_report():
    selected_date  = request.args.get("date", date.today().isoformat())
    student_filter = request.args.get("student_id", "")

    teacher_id = session.get("teacher_id")
    teacher_subjects = get_teacher_subjects(teacher_id) if teacher_id else []
    subject_names = [s['name'] for s in teacher_subjects]
    
    records = get_attendance_records_for_report(selected_date, student_filter, subjects=subject_names)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Student ID", "Subject", "Hour", "Date", "Time", "Status", "Auth Method"])
    for r in records:
        writer.writerow([
            r.get("student_id"),
            r.get("subject", ""),
            r.get("hour", ""),
            r.get("date"),
            r.get("marked_at", "")[:19] if r.get("marked_at") else "",
            r.get("status"),
            r.get("auth_method")
        ])

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=attendance_{selected_date}.csv"}
    )


def get_attendance_records_for_report(selected_date, student_filter="", subjects=None):
    """Helper: fetch attendance records filtered by date, student, and subjects."""
    all_records = get_attendance_records(
        student_id=student_filter if student_filter else None,
        subjects=subjects
    )
    return [r for r in all_records if r.get("date") == selected_date]



# ── Hour slot label map ────────────────────────────────────────────────────────
HOUR_LABELS = {
    "1": "7:30AM-9:00AM",
    "2": "9:45AM-10:45AM",
    "3": "10:45AM-11:45AM",
    "4": "11:45AM-12:45PM",
    "5": "12:45PM-1:45PM",
}


@app.route("/download-excel")
def download_excel():
    selected_date = request.args.get("date", date.today().isoformat())
    student_filter = request.args.get("student_id", "")
    hour_filter = request.args.get("hour", "all")
    mismatch_filter = request.args.get("status", "")

    teacher_id = session.get("teacher_id")
    teacher_subjects = get_teacher_subjects(teacher_id) if teacher_id else []
    subject_names = [s['name'] for s in teacher_subjects]
    
    all_students = get_all_students()
    records = get_attendance_records_for_report(selected_date, student_filter, subjects=subject_names)

    # Filter by hour slot
    if hour_filter != "all":
        try:
            h = int(hour_filter)
            records = [r for r in records if str(r.get("hour", "")) == str(h)]
        except ValueError:
            pass

    # Apply special mismatch filters
    if mismatch_filter:
        day_records = get_attendance_records_for_report(selected_date, "", subjects=subject_names)
        student_hours = {}
        for r in day_records:
            sid = r.get("student_id")
            h = str(r.get("hour", ""))
            if sid not in student_hours:
                student_hours[sid] = set()
            if h:
                student_hours[sid].add(h)
        
        target_student_ids = []
        for s in all_students:
            sid = s["id"]
            hours = student_hours.get(sid, set())
            has_h1 = "1" in hours
            has_h2 = "2" in hours
            
            if mismatch_filter == "missed1_att2":
                if not has_h1 and has_h2:
                    target_student_ids.append(sid)
            elif mismatch_filter == "att1_missed2":
                if has_h1 and not has_h2:
                    target_student_ids.append(sid)
                    
        records = [r for r in records if r.get("student_id") in target_student_ids]


    # ── Build workbook ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active

    hour_label = HOUR_LABELS.get(hour_filter, "All Hours") if hour_filter != "all" else "All Hours"
    ws.title   = "Attendance"

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value     = f"Attendance Report — {selected_date} | {hour_label}"
    title_cell.font      = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor="1E3A8A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["Student ID", "Subject", "Hour", "Time Slot", "Date", "Time", "Status", "Auth Method"]
    ws.merge_cells("A1:H1")           # re-merge to span 8 columns
    for col, h in enumerate(headers, start=1):
        cell            = ws.cell(row=2, column=col, value=h)
        cell.font       = Font(bold=True, color="FFFFFF")
        cell.fill       = PatternFill("solid", fgColor="2563EB")
        cell.alignment  = Alignment(horizontal="center")

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, r in enumerate(records, start=3):
        hour_num  = str(r.get("hour", ""))
        slot_lbl  = HOUR_LABELS.get(hour_num, f"Hour {hour_num}")
        marked_at = r.get("marked_at", "") or ""
        date_val  = r.get("date", "")
        time_val  = marked_at[11:19] if len(marked_at) >= 19 else (marked_at[:16] if marked_at else "")

        row_data = [
            r.get("student_id", ""),
            r.get("subject", ""),
            f"Hour {hour_num}" if hour_num else "—",
            slot_lbl,
            date_val,
            time_val,
            r.get("status", ""),
            r.get("auth_method", ""),
        ]
        for col, val in enumerate(row_data, start=1):
            cell            = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment  = Alignment(horizontal="left")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EFF6FF")  # light blue stripe

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [16, 22, 10, 24, 14, 12, 12, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Stream response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_hour = hour_filter.replace("/", "-")
    filename  = f"attendance_{selected_date}_hour{safe_hour}.xlsx"

    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )




# ══════════════════════════════════════════════════════════════════════════════
#  STUDENT ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/student/login", methods=["GET", "POST"])
def student_login():
    if request.method == "POST":
        password   = request.form.get("password", "").strip()

        student_id = request.form.get("student_id", "").strip()
        if not student_id or not password:
            flash("Student ID and Password are required.", "error")
            return redirect(url_for("student_login"))

        student = get_student(student_id)
        if student and student.get('password_hash'):
            if check_password_hash(student['password_hash'], password):
                session["student_id"] = student_id
                session["student_email"] = student['email']
                return redirect(url_for("student_dashboard"))
            else:
                flash("Invalid password.", "error")
        elif student:
            flash("This account requires a password update. Please register again.", "warning")
        else:
            flash("Student ID not found.", "error")
            
        return redirect(url_for("student_login"))

    return render_template("student_login.html")


@app.route("/student/register", methods=["GET", "POST"])
def student_register():
    if request.method == "POST":
        password   = request.form.get("password", "").strip()

        student_id = request.form.get("student_id", "").strip()
        name       = request.form.get("name", "").strip()
        email      = request.form.get("email", "").strip()
        if not student_id or not name or not email or not password:
            flash("All fields including password are required.", "error")
            return redirect(url_for("student_register"))

        student = get_student(student_id)
        if student and student.get('password_hash'):
            flash("Student ID already registered. Please login.", "warning")
            return redirect(url_for("student_login"))

        password_hash = generate_password_hash(password)
        if add_student(student_id, name):
            from database import update_student_password
            update_student_email(student_id, email)
            update_student_password(student_id, password_hash)
            session["student_id"] = student_id
            session["student_email"] = email
            flash(f"Welcome, {name}! Registration successful.", "success")
            return redirect(url_for("student_dashboard"))
        else:
            flash("Registration failed. Please try again.", "error")
            return redirect(url_for("student_register"))

    return render_template("student_login.html", register=True)


@app.route("/student/dashboard")
def student_dashboard():
    if "student_id" not in session:
        return redirect(url_for("student_login"))

    student_id   = session["student_id"]
    face_enrolled = check_face_enrolled(student_id)

    active_sessions = get_active_sessions()
    requested_session_id = request.args.get('request_verification', type=int)
    
    # Ensure requested session is in the list even if it just became inactive
    session_ids = [s["id"] for s in active_sessions]
    if requested_session_id and requested_session_id not in session_ids:
        req_session = get_session(requested_session_id)
        if req_session:
            active_sessions.insert(0, req_session)

    sessions_with_status = []
    for s in active_sessions:
        att_status = check_attendance_for_session(student_id, s["id"])
        sessions_with_status.append({
            "id":            s["id"],
            "subject":       s["subject"],
            "hour":          s["hour"],
            "date":          s["date"],
            "latitude":      s["latitude"],
            "longitude":     s["longitude"],
            "already_marked": att_status is not None,
            "status":        att_status,
            "otp_verified":  session.get(f"otp_verified_{s['id']}", False),
            "is_active":     s.get("is_active", 1) # Support checking if it's the requested but inactive one
        })

    my_records = get_attendance_records(student_id=student_id)

    return render_template(
        "student_dashboard.html",
        student_id=student_id,
        face_enrolled=face_enrolled,
        sessions=sessions_with_status,
        my_records=my_records[:10],
        classroom_lat=CLASSROOM_LAT,
        classroom_lon=CLASSROOM_LON,
        classroom_radius=CLASSROOM_RADIUS
    )



@app.route("/student/enroll-face", methods=["POST"])
def enroll_face():
    global recognizer, label_map

    if "student_id" not in session:
        flash("Login required.", "error")
        return redirect(url_for("student_login"))

    student_id = session["student_id"]
    
    # Try browser-side capture data first
    face_base64 = request.form.get("face_image")
    if face_base64:
        face = decode_base64_image(face_base64)
    else:
        # Fallback for local development (server-side camera)
        face = capture_face()

    if face is None:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest" or face_base64:
            return jsonify({"success": False, "message": "No face detected in scan."})
        flash("Face capture cancelled or no face detected.", "error")
        return redirect(url_for("student_dashboard"))

    if save_face(student_id, face):
        recognizer, label_map = train_recognizer()
        if face_base64:
            flash("Face scan added successfully!", "success")
            return jsonify({"success": True, "message": "Face scan added!"})
        flash("Face scan added! Add 3-5 scans from different angles/lighting for best accuracy.", "success")
    else:
        flash("Failed to save face. Please try again.", "error")

    return redirect(url_for("student_dashboard"))


@app.route("/student/mark-attendance/<int:session_id>", methods=["POST"])
def mark_attendance_route(session_id):
    global recognizer, label_map

    if "student_id" not in session:
        return redirect(url_for("student_login"))

    student_id = session["student_id"]

    # Validate session is still active
    att_session = get_session(session_id)
    if not att_session or not att_session["is_active"]:
        flash("This attendance session is no longer active.", "error")
        return redirect(url_for("student_dashboard"))

    # Duplicate check
    if check_attendance_for_session(student_id, session_id):
        flash("Attendance already marked for this session!", "warning")
        return redirect(url_for("student_dashboard"))

    # Final Step: Location Verification (Geofencing)
    lat_str = request.form.get("latitude")
    lon_str = request.form.get("longitude")

    if not lat_str or not lon_str:
        flash("Location access is required to mark attendance.", "error")
        return redirect(url_for("student_dashboard"))

    try:
        student_lat = float(lat_str)
        student_lon = float(lon_str)
        
        target_lat = att_session.get("latitude") if att_session.get("latitude") is not None else CLASSROOM_LAT
        target_lon = att_session.get("longitude") if att_session.get("longitude") is not None else CLASSROOM_LON
        
        distance = haversine_distance(student_lat, student_lon, target_lat, target_lon)
        
        if distance > CLASSROOM_RADIUS:
            flash(f"Out of range ({distance:.1f}m). Please go inside to mark attendance.", "error")
            return redirect(url_for("student_dashboard"))
            
        # STEP 2: Persistent OTP Verification
        qr_token = request.form.get("qr_token")
        otp_verified_in_session = session.get(f"otp_verified_{session_id}")
        
        # If not verified in current session, try to verify the token provided
        if not otp_verified_in_session:
            if qr_token and verify_otp(session_id, qr_token):
                session[f"otp_verified_{session_id}"] = True
            else:
                flash("OTP verification failed or expired. Please enter the current code from the teacher's screen.", "error")
                return redirect(url_for("student_dashboard", request_verification=session_id))

        # STEP 3: Face Recognition Flow
        from face_manager import capture_face, train_recognizer, recognize_face
        
        # Try browser-side capture data first
        face_base64 = request.form.get("face_image")
        if face_base64:
            face = decode_base64_image(face_base64)
        else:
            # Fallback for local development (server-side camera)
            face = capture_face()

        if face is None:
            flash("Face capture failed or no face detected. Please try again.", "error")
            return redirect(url_for("student_dashboard"))

        # Train recognizer if not ready
        if recognizer is None:
            recognizer, label_map = train_recognizer()

        # Recognize
        if recognizer is not None and label_map:
            matched_id, confidence = recognize_face(recognizer, label_map, face)
            print(f"[DEBUG] Face Recognition - Student ID: {student_id}, Predicted ID: {matched_id}, Confidence: {confidence:.2f}")
        else:
            print("[DEBUG] Face Recognition - Model not ready, retraining...")
            recognizer, label_map = train_recognizer()
            if recognizer and label_map:
                matched_id, confidence = recognize_face(recognizer, label_map, face)
            else:
                flash("Face recognition model is not initialized. Please enroll your face.", "error")
                return redirect(url_for("student_dashboard"))

        # Handle Match
        if matched_id:
            if str(matched_id) == str(student_id):
                if mark_attendance(student_id, session_id, auth_method="Face", confidence_score=round(confidence, 2)):
                    flash(f"Face recognized! Attendance marked for {att_session['subject']}.", "success")
                    return redirect(url_for("attendance_confirmed", session_id=session_id, method="Face"))
            else:
                # Recognized as someone else
                flash(f"Face recognized as a different student. Please try again or request manual verification.", "warning")
                print(f"[FACE] Mismatch: Student {student_id} scanned as {matched_id} (conf: {confidence:.2f})")
                return redirect(url_for("student_dashboard", request_verification=session_id))
        
        # No Match at all
        flash("Face not recognized. Tip: Add more face scans from different angles to improve accuracy.", "warning")
        return redirect(url_for("student_dashboard", request_verification=session_id))
            
    except ValueError:
        flash("Invalid location data received.", "error")
        return redirect(url_for("student_dashboard"))


@app.route("/student/request-verification/<int:session_id>", methods=["POST"])
def student_request_verification(session_id):
    if "student_id" not in session:
        return redirect(url_for("student_login"))
    
    student_id = session["student_id"]
    att_session = get_session(session_id)
    if not att_session or not att_session.get("is_active"):
        flash("Session not found or inactive.", "error")
        return redirect(url_for("student_dashboard"))

    # 1. Location Verification
    lat_str = request.form.get("latitude")
    lon_str = request.form.get("longitude")
    if not lat_str or not lon_str:
        flash("Location access is required for manual verification.", "error")
        return redirect(url_for("student_dashboard"))

    try:
        student_lat = float(lat_str)
        student_lon = float(lon_str)
        target_lat = att_session.get("latitude") or CLASSROOM_LAT
        target_lon = att_session.get("longitude") or CLASSROOM_LON
        distance = haversine_distance(student_lat, student_lon, target_lat, target_lon)
        
        if distance > CLASSROOM_RADIUS:
            flash(f"Out of range ({distance:.1f}m). You must be in the classroom.", "error")
            return redirect(url_for("student_dashboard"))
    except ValueError:
        flash("Invalid location data.", "error")
        return redirect(url_for("student_dashboard"))

    # 2. OTP Verification (Check session first)
    otp = request.form.get("otp")
    if not session.get(f"otp_verified_{session_id}"):
        if otp and verify_otp(session_id, otp):
             session[f"otp_verified_{session_id}"] = True
        else:
            flash("Invalid or expired OTP. Manual verification requires the current code.", "error")
            return redirect(url_for("student_dashboard", request_verification=session_id))

    # 3. Create pending record
    if mark_attendance(student_id, session_id, auth_method="Manual", 
                        confidence_score=None, status="Pending Verification"):
        flash("Verification request sent! Please wait for your teacher to approve.", "success")
    else:
        flash("Request already sent or attendance already marked.", "warning")
        
    return redirect(url_for("student_dashboard"))




@app.route("/student/confirmed")
def attendance_confirmed():
    if "student_id" not in session:
        return redirect(url_for("student_login"))

    student_id  = session["student_id"]
    method      = request.args.get("method", "Unknown")
    session_id  = request.args.get("session_id")

    att_session = None
    if session_id:
        att_session = get_session(int(session_id))

    return render_template("confirmed.html",
                           student_id=student_id,
                           auth_method=method,
                           att_session=att_session)


@app.route("/student/my-attendance")
def student_my_attendance():
    if "student_id" not in session:
        return redirect(url_for("student_login"))

    student_id = session["student_id"]
    summary    = get_student_subject_summary(student_id)
    all_records = get_attendance_records(student_id=student_id)

    return render_template(
        "student_my_attendance.html",
        student_id=student_id,
        subjects=summary['subjects'],
        overall=summary['overall'],
        all_records=all_records
    )


@app.route("/student/logout")
def student_logout():
    session.clear()
    return redirect(url_for("student_login"))


@app.route("/session/<int:session_id>/otp")
def session_otp(session_id):
    """Endpoint for teacher dashboard to get current dynamic 6-digit OTP."""
    if "teacher_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    token = get_dynamic_otp(session_id)
    # Also return time remaining in current 10s slot
    time_remaining = 10 - (int(time.time()) % 10)
    
    return jsonify({
        "token": token,
        "expires_in": time_remaining
    })

@app.route("/student/mark-attendance-qr", methods=["POST"])
def mark_attendance_qr():
    if "student_id" not in session:
        return jsonify({"success": False, "message": "Login required"}), 401
    
    student_id = session["student_id"]
    session_id = request.form.get("session_id")
    scanned_token = request.form.get("token")
    lat_str = request.form.get("latitude")
    lon_str = request.form.get("longitude")

    if not all([session_id, scanned_token, lat_str, lon_str]):
        return jsonify({"success": False, "message": "Missing required fields"}), 400

    att_session = get_session(session_id)
    if not att_session or not att_session.get("is_active"):
        return jsonify({"success": False, "message": "Session is inactive"}), 404

    # 1. Location Check
    try:
        student_lat = float(lat_str)
        student_lon = float(lon_str)
        target_lat = att_session.get("latitude") or CLASSROOM_LAT
        target_lon = att_session.get("longitude") or CLASSROOM_LON
        distance = haversine_distance(student_lat, student_lon, target_lat, target_lon)
        
        if distance > CLASSROOM_RADIUS:
            return jsonify({"success": False, "message": f"Out of range ({distance:.1f}m)."}), 403
    except ValueError:
        return jsonify({"success": False, "message": "Invalid coordinates"}), 400

    # 2. Token Check (Allow +/- 1 slot for sync issues)
    valid_tokens = []
    current_slot = int(time.time() // 10)
    for slot in [current_slot, current_slot - 1]:
        token_input = f"{session_id}{app.secret_key}{slot}"
        valid_tokens.append(hashlib.sha256(token_input.encode()).hexdigest()[:10].upper())

    if scanned_token.upper() not in valid_tokens:
        return jsonify({"success": False, "message": "QR Code expired or invalid. Please scan the latest code."}), 403

    # 3. Mark Attendance
    if check_attendance_for_session(student_id, session_id):
        return jsonify({"success": False, "message": "Already marked!"}), 400

    if mark_attendance(student_id, session_id, auth_method="QR"):
        return jsonify({"success": True, "message": "Attendance marked successfully via QR!"})
    
    return jsonify({"success": False, "message": "Failed to mark attendance."}), 500


@app.route("/student/verify-otp-only", methods=["POST"])
def verify_otp_only():
    """Standalone OTP verification for the 'Step 1' UI feedback. Sets persistent session flag."""
    if "student_id" not in session:
        return jsonify({"success": False, "message": "Login required"}), 401
    
    session_id = request.form.get("session_id", type=int)
    token = request.form.get("token")
    
    if verify_otp(session_id, token):
        session[f"otp_verified_{session_id}"] = True
        return jsonify({"success": True, "message": "OTP Verified! Move to Proximity Check."})
    else:
        return jsonify({"success": False, "message": "Invalid or expired OTP. Try again."})

# ── LEAVE MANAGEMENT ROUTES ───────────────────────────────────────────────────

@app.route("/student/leaves")
def student_leaves():
    if "student_id" not in session:
        return redirect(url_for("student_login"))
    student_id = session["student_id"]
    from database import get_leave_requests
    my_leaves = get_leave_requests(student_id=student_id)
    return render_template("student_leaves.html", student_id=student_id, my_leaves=my_leaves)

@app.route("/teacher/manage-leaves")
def teacher_manage_leaves():
    if "teacher_id" not in session or session["teacher_id"] != "T1001":
        flash("Unauthorized: Only Dr. Sridevi (T1001) can manage leaves.", "error")
        return redirect(url_for("teacher_dashboard"))
    
    teacher_id = session["teacher_id"]
    teacher = get_teacher(teacher_id) or DEFAULT_TEACHER
    from database import get_leave_requests
    leave_requests = get_leave_requests()
    return render_template("teacher_leaves.html", teacher=teacher, leave_requests=leave_requests)

@app.route("/student/apply-leave", methods=["POST"])
def apply_leave():
    """Student applies for a leave (Medical, OD, Other)."""
    if "student_id" not in session:
        flash("Login required to apply for leave.", "error")
        return redirect(url_for("student_login"))
        
    student_id = session["student_id"]
    from_date  = request.form.get("from_date")
    to_date    = request.form.get("to_date")
    reason     = request.form.get("reason")
    other      = request.form.get("other_reason")
    
    if not from_date or not to_date or not reason:
        flash("Please provide all required leave details.", "warning")
        return redirect(url_for("student_dashboard"))
        
    if from_date > to_date:
        flash("Invalid date range: 'From Date' must be before or equal to 'To Date'.", "error")
        return redirect(url_for("student_leaves"))
        
    if create_leave_request(student_id, from_date, to_date, reason, other):
        flash("Leave request submitted successfully! Pending approval from Dr. Sridevi.", "success")
    else:
        flash("Failed to submit leave request. Please try again.", "error")
        
    return redirect(url_for("student_leaves"))

@app.route("/teacher/leave-action/<int:req_id>/<action>", methods=["POST"])
def teacher_leave_action(req_id, action):
    """Dr. Sridevi (T1001) approves or rejects a leave request."""
    if "teacher_id" not in session or session["teacher_id"] != "T1001":
        flash("Unauthorized: Only Dr. Sridevi (T1001) can manage leaves.", "error")
        return redirect(url_for("teacher_dashboard"))
        
    new_status = "Approved" if action == "approve" else "Rejected"
    if update_leave_status(req_id, new_status):
        flash(f"Leave request {new_status}!", "success")
    else:
        flash("Failed to update leave status.", "error")
        
    return redirect(url_for("teacher_manage_leaves"))

# ══════════════════════════════════════════════════════════════════════════════
#  RUN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=True)
